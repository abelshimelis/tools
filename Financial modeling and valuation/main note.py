import pandas as pd
from openpyxl import load_workbook
from input_handler import InputHandler
from financial_model import FinancialModel

def main():
    # Step 1 — InputHandler reads the Excel file
    handler = InputHandler("3 statement modeling template.xlsx")
    statements  = handler.get_statements()
    assumptions = handler.get_assumptions()
    owc         = handler.get_owc()
    depreciation= handler.get_depreciation_schedule()
    debt        = handler.get_debt_schedule()

    print("Depreciation Schedule:")
    print(depreciation)
    print("\nOWC Schedule:")
    print(owc)
    print("\nDebt Schedule:")
    print(debt)

    print("Historical Income Statement (full):")
    income_full = pd.read_excel("3 statement modeling template.xlsx", sheet_name="Income Statement", header=1, index_col=0)
    print(income_full[['2022A', '2023A', '2024A']].head(20))
    print("\nAssumptions:")
    print(assumptions)

    # Step 2 — Pass everything into FinancialModel
    model = FinancialModel(statements, assumptions, owc, depreciation, debt)
    projections = model.project()

    # Print 2026P to 2029P Income Statement Projections
    print("Income Statement Projections (2025P to 2029P):")
    income_keys = ["Net Sales", "Membership Income", "Total Revenue", "COGS", "Gross Profit", "SGA", "EBITDA", "EBIT", "Prepaid", "Net Income"]
    for year in ['2025P', '2026P', '2027P', '2028P', '2029P']:
        print(f"\n{year}:")
        for key in income_keys:
            if key in projections.columns:
                value = projections.loc[year, key]
                print(f"  {key}: {value}")

    # Print Balance Sheet Projections
    print("\n\nBalance Sheet Projections (2025P to 2029P):")
    balance_keys = ["Ending Cash", "AR", "Inventory", "Prepaid", "PPE Net", "Goodwill", "Other Assets", "Total Assets", "ST Debt End", "AP", "Accrued Liab", "Accrued Tax", "LT Debt", "Share Capital", "Retained Earnings", "Total Equity"]
    for year in ['2025P', '2026P', '2027P', '2028P', '2029P']:
        print(f"\n{year}:")
        for key in balance_keys:
            if key in projections.columns:
                value = projections.loc[year, key]
                print(f"  {key}: {value}")

    # Print Cash Flow Statement Projections
    print("\n\nCash Flow Statement Projections (2025P to 2029P):")
    cashflow_keys = ["Net Income", "DA", "Deferred Tax", "Other Oper", "Net OWC Change", "CFO", "CAPEX CF", "Investments", "CFI", "ST Debt Chg", "LT Debt Chg", "Dividends", "Share Issuance", "CFF", "Net Chg Cash", "Ending Cash", "Free Cash Flow"]
    for year in ['2025P', '2026P', '2027P', '2028P', '2029P']:
        print(f"\n{year}:")
        for key in cashflow_keys:
            if key in projections.columns:
                value = projections.loc[year, key]
                print(f"  {key}: {value}")

    # Step 3 — Write projections back to a template-based workbook
    output_path = "3 statement modeling template output_new.xlsx"
    base_path = "3 statement modeling template.xlsx"

    # Read all existing sheets from the template to preserve structure
    existing_sheets = pd.read_excel(base_path, sheet_name=None)

    # Define specific projection sections (map model keys to close template labels)
    income_map = {
        "Net Sales": "Net sales",
        "Membership Income": "Membership and other income",
        "Total Revenue": "Total Revenue ($mm)",
        "COGS": "Cost of goods sold",
        "Gross Profit": "Gross Profit",
        "SGA": "SG&A",
        "EBITDA": "EBITDA",
        "EBIT": "EBIT",
        "Prepaid": "Prepaid expenses and other",
        "Net Income": "Net income"
    }

    balance_map = {
        "Ending Cash": "Cash and cash equivalents",
        "AR": "Receivables, net",
        "Inventory": "Inventories",
        "Prepaid": "Prepaid expenses and other",
        "PPE Net": "Property, plant & equipment, net",
        "Goodwill": "Goodwill",
        "Other Assets": "Other assets & deferred charges",
        "Total Assets": "Total assets",
        "ST Debt End": "Short-term borrowings",
        "AP": "Accounts payable",
        "Accrued Liab": "Accrued liabilities",
        "Accrued Tax": "Accrued income taxes",
        "LT Debt": "Long-term debt",
        "Share Capital": "Share capital & premium",
        "Retained Earnings": "Retained earnings",
        "Total Equity": "Total equity"
    }

    cashflow_map = {
        "Net Income": "Net income",
        "DA": "Depreciation & amortization (non-cash add-back)",
        "Deferred Tax": "Deferred income taxes (non-cash add-back)",
        "Other Oper": "Other operating activities",
        "Net OWC Change": "Changes in operating working capital",
        "CFO": "Cash from operating activities",
        "CAPEX CF": "Capital expenditures",
        "Investments": "Investments & acquisitions",
        "CFI": "Cash from investing activities",
        "ST Debt Chg": "Net change in short-term borrowings",
        "LT Debt Chg": "Net change in long-term debt",
        "Dividends": "Dividends",
        "Share Issuance": "Share issuance / (buyback)",
        "CFF": "Cash from financing activities",
        "Net Chg Cash": "Net change in cash",
        "Ending Cash": "Ending cash balance",
        "Free Cash Flow": "Free cash flow"
    }

    def map_statement(proj_df, key_map):
        keys = [k for k in key_map.keys() if k in proj_df.columns]
        df = proj_df[keys].rename(columns={k: key_map[k] for k in keys})
        return df

    income_proj = map_statement(projections, income_map)
    balance_proj = map_statement(projections, balance_map)
    cashflow_proj = map_statement(projections, cashflow_map)

    print("\nIncome Statement Mapping:")
    for model_key, template_label in income_map.items():
        print(f"{model_key} -> {template_label}")

    wb = load_workbook(base_path)

    def find_year_columns(ws, target_years):
        year_cols = {}
        for r in range(1, min(20, ws.max_row) + 1):
            for c in range(1, ws.max_column + 1):
                kval = ws.cell(row=r, column=c).value
                if kval is None:
                    continue
                kval = str(kval).strip()
                if kval in target_years:
                    year_cols[kval] = c
            if len(year_cols) >= len(target_years):
                break
        return year_cols

    def write_projections_into_sheet(sheet_name, mapping):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]

        def set_cell_value(row, col, value):
            # openpyxl cannot write directly to MergedCell objects, so write to top-left of merged range if needed
            coord = ws.cell(row=row, column=col).coordinate
            for merged in ws.merged_cells.ranges:
                if coord in merged:
                    ws.cell(row=merged.min_row, column=merged.min_col).value = value
                    return
            ws.cell(row=row, column=col).value = value

        years = [str(y) for y in projections.index.tolist()]
        year_cols = find_year_columns(ws, years)
        if not year_cols:
            return

        row_map = {}
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is None:
                continue
            val_str = str(val).strip()
            for template_label in mapping.values():
                if val_str == template_label:
                    row_map[template_label] = r
                    break

        print(f"\nFound labels in {sheet_name}:")
        for label, row in row_map.items():
            print(f"Row {row}: {label}")

        for model_key, template_label in mapping.items():
            if template_label not in row_map:
                continue
            if model_key not in projections.columns:
                continue
            r = row_map[template_label]
            for year in years:
                if year in year_cols:
                    set_cell_value(r, year_cols[year], float(projections.loc[year, model_key]))

    write_projections_into_sheet("Income Statement", income_map)
    write_projections_into_sheet("Balance Sheet", balance_map)
    write_projections_into_sheet("Cash Flow Statement", cashflow_map)

    # Do not create extra projection sheets. Data is written directly into existing statement sheets.
    wb.save(output_path)

    print(f"Projection workbook saved to '{output_path}'")
    print("Projection sample (Income Statement):")
    print(income_proj.head())

if __name__ == "__main__":
    main()