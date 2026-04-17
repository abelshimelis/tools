import os
import sys
import pandas as pd
from openpyxl import load_workbook

# Makes sure exe finds files in the right place when compiled with PyInstaller
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_PATH = os.path.join(BASE_DIR, "Bank_Financial_Model_5Y.xlsx")

from bank_input_handler import InputHandler
from bank_financial_model import FinancialModel


def find_year_columns(ws, target_years):
    year_cols = {}
    for r in range(1, min(20, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            kval = ws.cell(row=r, column=c).value
            if kval is None:
                continue
            kval = str(kval).strip()
            if kval in target_years:
                year_cols[kval] = c
        if len(year_cols) >= len(target_years):
            break
    return year_cols


def set_cell_value(ws, row, col, value):
    coord = ws.cell(row=row, column=col).coordinate
    for merged in ws.merged_cells.ranges:
        if coord in merged:
            ws.cell(row=merged.min_row, column=merged.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def write_projections_into_sheet(ws, mapping, proj_df):
    years = [str(y) for y in proj_df.index.tolist()]
    year_cols = find_year_columns(ws, years)
    if not year_cols:
        return

    row_map = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is None:
            continue
        val_str = str(val).strip()
        for model_key, template_label in mapping.items():
            if val_str == template_label:
                row_map[model_key] = r
                break

    for model_key, template_label in mapping.items():
        if template_label not in row_map:
            continue
        if model_key not in proj_df.columns:
            continue
        r = row_map[model_key]
        for year in years:
            if year in year_cols:
                set_cell_value(ws, r, year_cols[year], float(proj_df.loc[year, model_key]))


def main():
    print("=" * 50)
    print("Bank Financial Model — 3 Statement Projection")
    print("=" * 50)

    # ── Step 1: Read Excel ────────────────────────────────────────────
    print("\n[1/4] Reading Excel template...")
    try:
        handler = InputHandler(DATA_PATH)
        assumptions    = handler.get_assumptions()
        income_hist    = handler.get_income_statement()
        balance_hist   = handler.get_balance_sheet()
        capital_hist   = handler.get_capital_adequacy()
        cashflow_hist  = handler.get_cash_flow()
        print(f"      Done. Loaded: {DATA_PATH}")
    except FileNotFoundError:
        print(f"\nERROR: Could not find template at: {DATA_PATH}")
        print("Make sure the Excel file is in the same folder as bank_main.py.")
        sys.exit(1)
    except Exception as e:
        print(f"\nERROR reading Excel: {e}")
        sys.exit(1)

    # ── Step 2: Run projections ───────────────────────────────────────
    print("\n[2/4] Running projections...")
    try:
        model = FinancialModel(assumptions, income_hist, balance_hist, capital_hist, cashflow_hist)
        projections = model.project()
        print("      Done.")
    except Exception as e:
        print(f"\nERROR in projection calculations: {e}")
        sys.exit(1)

    # ── Step 2b: Write projections back into the template workbook ─────
    print("\n[3/4] Writing projections into template...")
    wb = load_workbook(DATA_PATH)

    income_map = {
        'Interest Income': 'Interest Income',
        'Interest Expense': 'Interest Expense',
        'Net Interest Income': 'Net Interest Income (NII)',
        'Fee & Commission': 'Fee & Commission Income',
        'Trading Income': 'Trading Income',
        'Other Non-Interest': 'Other Operating Income',
        'Total Non-Interest': 'Total Non-Interest Income',
        'Total Operating Income': 'Total Operating Income',
        'Total Opex': 'Total Operating Expenses',
        'Staff Costs': 'Staff Costs',
        'Tech & Infra': 'Technology & Infrastructure',
        'Occupancy & Other': 'Occupancy & Admin',
        'Depreciation': 'Depreciation & Amortisation',
        'Pre-Provision Profit': 'Pre-Provision Operating Profit (PPOP)',
        'Loan Loss Provisions': 'Loan Loss Provisions (LLP)',
        'Pre-Tax Profit': 'Pre-Tax Profit (PBT)',
        'Tax Expense': 'Income Tax Expense',
        'Net Profit After Tax': 'Net Income (PAT)',
        'NIM': 'NIM (%)',
        'Cost to Income Ratio': 'Cost-to-Income (%)',
        'LLP / Loans (%)': 'LLP / Loans (%)',
        'ROE': 'ROE (%)',
        'ROA': 'ROA (%)'
    }

    balance_map = {
        'Cash & Reserves': 'Cash & Central Bank Reserves',
        'Interbank Placements': 'Interbank Placements',
        'Investment Securities': 'Investment Securities',
        'Gross Loans': 'Gross Loans & Advances',
        'Loan Loss Reserves': 'Less: Loan Loss Reserves',
        'Net Loans': 'Net Loans & Advances',
        'Fixed Assets': 'Fixed Assets & ROU',
        'Goodwill': 'Goodwill & Intangibles',
        'Other Assets': 'Other Assets (Deferred Tax, Derivatives)',
        'Total Assets': 'TOTAL ASSETS',
        'CASA Deposits': 'CASA Deposits (Current & Savings)',
        'Term Deposits': 'Term / Fixed Deposits',
        'Total Deposits': 'Total Customer Deposits',
        'Interbank Borrowings': 'Interbank Borrowings',
        'Debt Securities': 'Debt Securities Issued',
        'Subordinated Debt': 'Subordinated Debt',
        'Other Liabilities': 'Other Liabilities (Derivatives, Accruals)',
        'Total Liabilities': 'TOTAL LIABILITIES',
        'Share Capital': 'Share Capital & Premium',
        'Retained Earnings': 'Retained Earnings',
        'OCI Reserves': 'Other Comprehensive Income (OCI)',
        'AT1 Capital': 'AT1 Capital Instruments',
        'Total Equity': 'TOTAL EQUITY',
        'Total Liab & Equity': 'TOTAL LIABILITIES & EQUITY',
        'LDR': 'Loan-to-Deposit Ratio (LDR %)',
        'CET1 Ratio': 'CET1 Ratio (%)',
        'NPL Ratio': 'NPL Ratio (%)'
    }

    capital_map = {
        'CET1 Capital': 'Common Equity Tier 1 (CET1)',
        'AT1 Capital': 'Additional Tier 1 (AT1)',
        'Tier2 Capital': 'Tier 2 Capital (Sub Debt + Reserves)',
        'Credit Risk RWA': 'Credit Risk RWA',
        'Market Risk RWA': 'Market Risk RWA',
        'Operational Risk RWA': 'Operational Risk RWA',
        'CET1 Ratio': 'CET1 Ratio (%)',
        'Tier1 Ratio': 'Tier 1 Ratio (%)',
        'CAR': 'Total Capital Adequacy Ratio (%)'
    }

    cashflow_map = {
        'Net Profit': 'Net Income',
        'Depreciation': 'Add: Depreciation & Amortisation',
        'Provisions Add-back': 'Add: Loan Loss Provisions',
        'Change in Loans': 'Change in Loans (net)',
        'Change in Deposits': 'Change in Deposits (net)',
        'Change in Other WC': 'Change in Other Working Capital',
        'CFO': 'Net Cash from Operating Activities',
        'Purchase of Securities': 'Purchase of Investment Securities',
        'Capex': 'Capital Expenditure (Capex)',
        'Acquisitions': 'Acquisitions / Disposals (net)',
        'CFI': 'Net Cash from Investing Activities',
        'Dividends': 'Dividends Paid',
        'Debt Issuance': 'Issuance of Debt Securities',
        'Share Buybacks': 'Share Capital Issued / (Bought Back)',
        'CFF': 'Net Cash from Financing Activities',
        'Net Change in Cash': 'NET CHANGE IN CASH'
    }

    write_projections_into_sheet(wb['Income Statement'], income_map, projections['income'])
    write_projections_into_sheet(wb['Balance Sheet'], balance_map, projections['balance'])
    write_projections_into_sheet(wb['Capital Adequacy'], capital_map, projections['capital'])
    write_projections_into_sheet(wb['Cash Flow'], cashflow_map, projections['cashflow'])

    wb.save(DATA_PATH)
    print(f"      Done. Overwrote template with projections at: {DATA_PATH}")

    # ── Step 3: Print results ─────────────────────────────────────────
    print("\n[4/4] Results:\n")

    print("--- INCOME STATEMENT ($mm) ---")
    income_display = projections['income'].filter([
        'Interest Income', 'Interest Expense', 'Net Interest Income',
        'Fee & Commission', 'Trading Income', 'Other Non-Interest',
        'Total Non-Interest', 'Total Operating Income', 'Total Opex',
        'Staff Costs', 'Tech & Infra', 'Occupancy & Other',
        'Depreciation', 'Pre-Provision Profit', 'Loan Loss Provisions',
        'Pre-Tax Profit', 'Tax Expense', 'Net Profit After Tax'
    ], axis=1)
    print(income_display.T.to_string())

    print("\n--- BALANCE SHEET ($mm) ---")
    balance_display = projections['balance'].filter([
        'Cash & Reserves', 'Interbank Placements', 'Investment Securities',
        'Gross Loans', 'Loan Loss Reserves', 'Net Loans', 'Fixed Assets',
        'Goodwill', 'Other Assets', 'Total Assets', 'CASA Deposits',
        'Term Deposits', 'Total Deposits', 'Interbank Borrowings',
        'Debt Securities', 'Subordinated Debt', 'Other Liabilities',
        'Total Liabilities', 'Share Capital', 'Retained Earnings',
        'OCI Reserves', 'AT1 Capital', 'Total Equity', 'Total Liab & Equity'
    ], axis=1)
    print(balance_display.T.to_string())

    print("\n--- CAPITAL ADEQUACY ---")
    cap_display = projections['capital'].filter([
        'CET1 Capital', 'AT1 Capital', 'Tier1 Capital', 'Tier2 Capital',
        'Total Reg Capital', 'Credit Risk RWA', 'Market Risk RWA',
        'Operational Risk RWA', 'Total RWA', 'CET1 Ratio', 'Tier1 Ratio',
        'CAR', 'LDR'
    ], axis=1)
    print(cap_display.T.to_string())

    print("\n--- CASH FLOW ($mm) ---")
    cf_display = projections['cashflow'].filter([
        'Net Profit', 'Depreciation', 'Provisions Add-back', 'Change in Loans',
        'Change in Deposits', 'Change in Other WC', 'CFO',
        'Purchase of Securities', 'Capex', 'Acquisitions', 'CFI',
        'Debt Issuance', 'Dividends', 'Share Buybacks', 'CFF',
        'Net Change in Cash'
    ], axis=1)
    print(cf_display.T.to_string())

    print("\n--- BALANCE SHEET CHECK ---")
    print(projections['balance']['BS Check'].to_string())

    output_path = os.path.join(BASE_DIR, "banks 3 statement output.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        projections['income'].to_excel(writer, sheet_name="Income Statement")
        projections['balance'].to_excel(writer, sheet_name="Balance Sheet")
        projections['capital'].to_excel(writer, sheet_name="Capital Adequacy")
        projections['cashflow'].to_excel(writer, sheet_name="Cash Flow")
    print(f"\n[4/4] Complete. Projections exported to: {output_path}")


if __name__ == "__main__":
    main()
